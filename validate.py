import tkinter as tk
from tkinter import filedialog
import openpyxl
import smtplib
import dns.resolver

def validate_email(email):
    # Проверяем наличие символа "@" в адресе электронной почты
    if '@' not in email:
        return False

    # Разделяем адрес электронной почты для извлечения домена
    domain = email.split('@')[1]

    try:
        # Запрашиваем MX-записи домена
        records = dns.resolver.resolve(domain, 'MX')
        mx_record = str(records[0].exchange)

        # Подключаемся к SMTP-серверу домена
        server = smtplib.SMTP()
        server.set_debuglevel(0)
        server.connect(mx_record)
        server.helo(server.local_hostname)
        server.mail('me@domain.com')

        # Проверяем код ответа для адреса электронной почты
        code, message = server.rcpt(str(email))
        server.quit()

        # Если код ответа равен 250, адрес электронной почты действителен
        if code == 250:
            return True
        else:
            return False
    except dns.resolver.NXDOMAIN:
        return False
    except smtplib.SMTPConnectError:
        return False
    except smtplib.SMTPServerDisconnected:
        return False
    except smtplib.SMTPResponseException:
        return False
    except:
        return False

def validate_emails():
    # Открываем диалоговое окно для выбора файла Excel
    filepath = filedialog.askopenfilename(title="Выберите файл Excel",
                                          filetypes=(("Файлы Excel", "*.xlsx"),
                                                     ("Все файлы", "*.*")))

    if filepath:
        # Загружаем файл Excel
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        # Создаем новую книгу Excel для сохранения действительных адресов электронной почты
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        valid_count = 0
        row_index = 1

        # Перебираем строки в файле Excel
        for row in sheet.iter_rows(values_only=True):
            email = row[0]

            if email and validate_email(email):
                # Если адрес электронной почты действителен, записываем его в новую книгу
                new_sheet.cell(row=row_index, column=1, value=email)
                valid_count += 1
                row_index += 1

            counter_value.set("Проверено: {}/{}".format(row_index-1, sheet.max_row-1))
            window.update()

        # Сохраняем новую книгу с действительными адресами электронной почты
        new_filepath = filedialog.asksaveasfilename(
            title="Сохранить действительные адреса электронной почты",
            defaultextension=".xlsx",
            filetypes=(("Файлы Excel", "*.xlsx"), ("Все файлы", "*.*")))
        if new_filepath:
            new_wb.save(new_filepath)

        # Обновляем метку счетчика
        counter_value.set("Действительные адреса электронной почты: {}".format(valid_count))
        window.update()

# Создаем окно Tkinter
window = tk.Tk()
window.title("Валидатор адресов электронной почты")

# Задаем размер и положение окна
window_width = 330
window_height = 130
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))
window.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Создаем кнопку для запуска процесса валидации адресов электронной почты
validate_button = tk.Button(window, text="Проверить адреса электронной почты",
                            command=validate_emails)
validate_button.pack(pady=10)

# Создаем метку для отображения счетчика
counter_value = tk.StringVar()
counter_label = tk.Label(window, textvariable=counter_value)
counter_label.pack()

# Запускаем цикл событий Tkinter
window.mainloop()
